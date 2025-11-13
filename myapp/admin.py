from typing import Any
from django.db.models.fields.related import ForeignKey
import openpyxl
from django import forms
from django.urls import reverse
from django.contrib import admin
from django.utils.html import format_html
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from django.utils.safestring import mark_safe
from django.contrib.auth import get_user_model 
from jalali_date.fields import JalaliDateField
from django.contrib.auth.models import Permission 
from django.contrib.auth.forms import UserChangeForm
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.http import  HttpRequest, HttpResponseRedirect , HttpResponse
from myapp.templatetags.custom_filters import to_jalali_persian , to_jalali
from .models import User ,ServiceRequests , ActiveLocations , ActiveModems , ActivePlans , OtherInfo , AccountNumber
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from rangefilter.filters import DateTimeRangeFilter
from jalali_date.admin import AdminJalaliDateWidget
from django.contrib import messages

class DisplayOnlyWidget(forms.Widget):
    def __init__(self, text, color="black", *args, **kwargs):
        self.text = text
        self.color = color
        super().__init__(*args, **kwargs)

    def render(self, name, value, attrs=None, renderer=None):
        html = f"<div style='padding:5px; color:{self.color}; font-weight:normal;'>{self.text}</div>"
        return mark_safe(html)

class NoorajamAdminSite(admin.AdminSite):
    site_header = 'پنل ادمین سایت نوراجم'
    site_title = 'پنل مدیریت'
    index_title = 'به پنل مدیریت شرکت توسعه زیرساخت نوراجم خوش آمدید'

    def get_app_list(self, request):
        app_list = super().get_app_list(request)
        for app in app_list:
            if app['app_label'] == 'myapp':
                custom_order = [
                    'ServiceRequests',
                    'ActiveLocations',
                    'ActivePlans',
                    'ActiveModems',
                    'OtherInfo',
                    'User',
                ]
                app['models'].sort(
                    key=lambda x: custom_order.index(x['object_name'])
                    if x['object_name'] in custom_order else len(custom_order)
                )
        return app_list

super_admin_site = NoorajamAdminSite(name='noorajam_admin')

@admin.register(ServiceRequests , site=super_admin_site)
class ServiceRequestsAdmin(admin.ModelAdmin):
    # list_display = ('first_name','last_name','location','mobile_number')
    list_display_links = ("full_name",)

    # list_filter = ("location","finalization_status","marketer_status","drop_status","fusion_status","pay_status","submission_status")
    search_fields = ("mobile_number","national_code","post_code")
    ordering = ("-request_time",)

    def get_list_display(self, request,obj=None):
        user = request.user
        roles = [
            user.role_supervisor,
            user.role_marketer,
            user.role_dropagent,
            user.role_fusionagent,
            user.role_operator,
        ]
        active_roles = sum(1 for r in roles if r)

        if user.is_superuser:
            return('full_name','ispan','mobile_number')

        if active_roles > 1:
            return ('full_name','location','mobile_number')
        else:
            return('full_name','ispan','mobile_number')

    def get_list_filter(self, request, obj=None):
        # ترتیب مرجع برای سوپریوزر
        superuser_order = [
            ("request_time",DateTimeRangeFilter),
            "location",
            "finalization_status",
            "marketer_status",
            "drop_status",
            "supervisor_status",
            "fusion_status",
            "pay_status",
            "submission_status",
        ]

        # اگر سوپریوزر است → لیست کامل و مرتب را بده
        if request.user.is_superuser:
            return tuple(superuser_order)

        # نقش‌ها و فیلدهای مرتبط با هر نقش
        role_fields = {
            "role_marketer": ("location","marketer_status", "pay_status"),
            "role_supervisor": ("location", "supervisor_status"),
            "role_dropagent": ("location", "drop_status"),
            "role_operator":("location","submission_status"),
            "role_fusionagent": ("location","fusion_status"),
        }

        # جمع‌آوری فیلدها بدون تکرار
        selected_fields = []

        for role_name, fields in role_fields.items():
            if getattr(request.user, role_name, False):
                for field in fields:
                    if field not in selected_fields:
                        selected_fields.append(field)

        # حالا فیلدها را بر اساس ترتیب مرجع مرتب کن
        ordered_fields = [f for f in superuser_order if f in selected_fields]

        return tuple(ordered_fields)

    def jalali_request_time(self, obj):
        return to_jalali_persian(obj.request_time)
    jalali_request_time.short_description = "زمان درخواست" # type: ignore

    def contact_user(self,obj):
        try:
            message_url = reverse('send_user_message',args=[obj.mobile_number])
            call_url = "tel:"+obj.mobile_number
        except Exception as e:
            print(e)
            message_url, call_url = "#" , "#"
        return format_html(f"""<a class="button" href="{message_url}">ارسال پیامک</a>  -  <a class="button" href="{call_url}">تماس</a>""")
    contact_user.short_description = "ارتباط با مشترک"        # type: ignore

    def download_form(self,obj):
        try:
            word_url = reverse('create-form',args=['word',obj.pk])
            pdf_url = reverse('create-form',args=['pdf',obj.pk])
        except:
            return format_html("-")

        return format_html(f"""<a class="button" href="{word_url}">Word دانلود</a>&nbsp; - 
                           <a class="button" href="{pdf_url}">Pdf دانلود</a>""")
    download_form.short_description = "دانلود فرم ثبت نام" # type: ignore

    class Media:
        css = {
            'all': ('css/admin_modal.css',)
        }
        js = ('js/admin_modal.js',) 

    def view_contract_button(self, obj):
        if obj.contract_snapshot:
            url = reverse('admin_contract_preview', args=[obj.id])
            return format_html(
                '<a class="button" href="javascript:void(0);" onclick="openContractModal(\'{}\')">📄 مشاهده قرارداد</a>',
                url
            )
        return "-"
    
    view_contract_button.short_description = "قرارداد"
    view_contract_button.allow_tags = True

    def documents_box(self, obj):
        if not obj.documents:
            return "فایلی آپلود نشده است."
            
        view_url = obj.documents.url
          
        download_url = reverse('download-document', args=[obj.pk])
        
        return format_html(
            f"""<a class="button" href="{download_url}">دانلود</a> 
                 - 
                <a class="button" href="{view_url}" target="_blank">مشاهده</a>"""
        )
    documents_box.short_description = "بخش مدارک"

    def full_name(self,obj):
        return obj.get_full_name()
    full_name.short_description = 'نام' # type: ignore
    
    def ispan(self,obj):
        user = getattr(self, '_current_user', None)
        if user:
            if user.is_superuser:
                if  obj.finalization_status == "ended":
                    return format_html('<span style="color:limegreen;">● تحویل و اتمام کار</span>')
                elif obj.submission_status == "registered":
                    return format_html('<span style="color: darkgoldenrod;">● در انتظار ثبت اتمام کار</span>')
                elif obj.submission_status != "registered":
                    return format_html('<span style="color: darkgoldenrod;">● سرویس درجریان نصب است</span>')

            if user.role_marketer:
                if obj.marketer_status == "accepted":
                    return format_html('<span style="color: limegreen;">● اطلاعات تایید شده</span>')
                elif obj.marketer_status == "rejected":
                    return format_html('<span style="color:red;">● اطلاعات رد شده</span>')
                elif obj.marketer_status == "pending":
                    return format_html('<span style="color:darkgoldenrod;">● در انتظار تایید اطلاعات</span>')
                
            if user.role_dropagent:
                if obj.drop_status == "accepted" :
                    return format_html('<span style="color: limegreen;">● دراپ کشی انجام شد</span>')
                elif obj.drop_status == "rejected":
                    return format_html('<span style="color:red;">● عدم امکان اجرای دراپ</span>')
                elif obj.drop_status == "pending":
                    return format_html('<span style="color:darkgoldenrod;">● در انتظار دراپ کشی</span>')
                elif obj.drop_status == "queued":
                    return format_html('<span style="color:darkgoldenrod;">● در صف دراپ کشی</span>')
                # elif obj.drop_status == "repending":
                #     return format_html('<span style="color:darkgoldenrod;">● دراپ توسط ناظر رد شد ، در انتظار بازبینی مجدد شما</span>')
                
            if user.role_supervisor:
                if obj.supervisor_status == "accepted":
                    return format_html('<span style="color:limegreen;">● دراپ و فیوژن تایید شد</span>')
                elif obj.supervisor_status == "rejected":
                    return format_html('<span style="color: red ;">● دراپ و فیوژن رد شد</span>')
                elif obj.supervisor_status == "pending":
                    return format_html('<span style="color: darkgoldenrod ;">● در انتظار بررسی ناظر</span>')
                
            if user.role_operator:
                if obj.submission_status == "registered":
                    return format_html('<span style="color:limegreen;">● فرم ثبت شده</span>')
                elif obj.submission_status == "pending":
                    return format_html('<span style="color: darkgoldenrod;">● در انتظار ثبت فرم</span>')

            
            if user.role_fusionagent:
                if obj.fusion_status == "accepted":
                    return format_html('<span style="color: limegreen;">● فیوژن زنی انجام شد</span>')
                elif obj.fusion_status == "pending":
                    return format_html('<span style="color: darkgoldenrod;">● در انتظار فیوژن زنی</span>')
                elif obj.fusion_status == "queued":
                    return format_html('<span style="color: darkgoldenrod;">● در صف فیوژن زنی</span>')
                elif obj.fusion_status== "repending":
                    return format_html('<span style="color:darkgoldenrod;">● نصب توسط ناظر رد شد ، در انتظار بازبینی مجدد شما</span>')
    ispan.short_description = 'وضعیت' # type: ignore

    def get_queryset(self, request):
        self._current_user = request.user
        from django.db.models import Q
        # ۱. دریافت کوئری‌ست پایه
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs

        combined_q = Q(pk=None) 
        
        user = request.user

        # شرط بازاریاب:
        if user.role_marketer:
            combined_q |= Q(drop_status=ServiceRequests.DropStatus.pending)

        # شرط دراپ کش:
        if user.role_dropagent:
            combined_q |= (
                Q(marketer_status=ServiceRequests.MarketerFormStatus.accepted) &
                Q(fusion_status=ServiceRequests.FusionStatus.pending)
            )

        # شرط فیوژن زن:
        if user.role_fusionagent:
            combined_q |= (
                Q(drop_status=ServiceRequests.DropStatus.accepted) &
                ~Q(supervisor_status=ServiceRequests.SuperVisorStatus.accepted)
            )

        # شرط ناظر (سوپروایزر):
        if user.role_supervisor:
            combined_q |= (
                ( Q(fusion_status=ServiceRequests.FusionStatus.accepted) | Q(fusion_status=ServiceRequests.FusionStatus.repending) ) &
                   Q(submission_status=ServiceRequests.SubmissionStatus.pending)
            )

        #شرط اپراتور
        if user.role_operator:
            combined_q |= (
                Q(supervisor_status=ServiceRequests.SuperVisorStatus.accepted) &
                Q(finalization_status=ServiceRequests.FinalizationStatus.pending)
            )      

        return qs.filter(combined_q).distinct()

    def has_add_permission(self, request, obj=None):
        return request.user.is_superuser

    def has_delete_permission(self, request, obj=None):
        return (True if request.user.is_superuser or request.user.role_marketer else False)

    def has_view_history_permission(self,request,obj=None): return request.user.is_superuser

    def formfield_for_choice_field(self, db_field, request, **kwargs):

        formfield = super().formfield_for_choice_field(db_field, request, **kwargs)

        if db_field.name == "fusion_status":
            all_choices = list(formfield.choices)

            # hidden choice
            hidden_value = "repending"

            # مقدار فعلی آبجکت رو پیدا کن
            obj_id = request.resolver_match.kwargs.get("object_id")
            current_value = None
            if obj_id:
                obj = ServiceRequests.objects.filter(pk=obj_id).first()
                if obj:
                    current_value = obj.fusion_status

            # اگر مقدار فعلی repending نیست → حذفش کن
            if current_value != hidden_value:
                formfield.choices = [c for c in all_choices if c[0] != hidden_value]

        return formfield

    def formfield_for_dbfield(self, db_field, request, **kwargs):
        if db_field.name == 'payment_date':
            kwargs['form_class'] = JalaliDateField
            kwargs['widget'] = AdminJalaliDateWidget
        return super().formfield_for_dbfield(db_field, request, **kwargs)

    def formfield_for_foreignkey(self,db_field,request,**kwargs):
        if db_field.name == "marketer_name":
            kwargs["queryset"] = User.objects.filter(role_marketer=True)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)
    
    def save_model(self, request, obj, form, change):
        if not change:
            pass

        elif change:
            old_obj = self.model.objects.get(pk=obj.pk)

            # اگر از pending به accepted رفت
            if old_obj.fusion_status in ['pending','repending','queued'] and form.cleaned_data.get('fusion_status','') == "accepted" :
                obj.supervisor_status = 'pending'

            # اگر ناظر رد کرد → وضعیت repending
            if old_obj.supervisor_status != 'rejected' and form.cleaned_data.get('supervisor_status','') == 'rejected':
                obj.fusion_status = 'repending'


            obj.save()
        super().save_model(request, obj, form, change)

    def get_fieldsets(self, request, obj=None):
        from collections import defaultdict

        MASTER_FIELDSETS_STRUCTURE = (
            ('وضعیت نصب', {
                'fields': ('marketer_status', 'drop_status', 'fusion_status','supervisor_status',
                            'submission_status','virtual_number','port_number','finalization_status'),
                'classes': ('collapse',)
            }),
            ('وضعیت پرداخت', {
                'fields': ('pay_status', 'account_number', 'tracking_payment','payment_date','payment_time'),
                'classes': ('collapse',)
            }),
            ('اطلاعات دراپ', {
                'fields': ('outdoor_area','internal_area','fat_index',
                           'odc_index','pole_count','headpole_count',
                           'hook_count'),
                'classes': ('collapse',)
            }),
            ('باکس دانلود', {
                'fields': ('documents_box','download_form','documents','view_contract_button','marketer_name'), 
                'classes': ('collapse',)
            }),
            ('اطلاعات شخصی', {
                'fields': (
                    'first_name', 'last_name', 'father_name', 'national_code','originated_from',
                    'bc_number', 'birthday','landline_number','mobile_number',
                    'address','location','post_code','house_is_owner',
                ),
                'classes': ('collapse',)
            }),
            ('اطلاعات سرویس', {
                'fields': ('sip_phone', 'modem','plan'),
                'classes': ('collapse',)
            }),
            ("سایر اطلاعات", {
                'fields': (
                    'finished_request','contact_user','tracking_code','log_msg_status',
                    'jalali_request_time','ip_address'
                ),
                'classes': ('collapse',)
            }),
        )

        # اگر ابرکاربر بود، لیست کامل را برگردان
        if request.user.is_superuser:
            return MASTER_FIELDSETS_STRUCTURE

        # --- ۲. تعریف دسترسی‌های هر نقش ---
        ROLE_FIELDS = {
            'role_marketer': {
                'وضعیت نصب': ['marketer_status','pay_status'],
                'وضعیت پرداخت':['pay_status', 'account_number', 'tracking_payment','payment_date','payment_time'],
                'اطلاعات دراپ': ['outdoor_area', 'internal_area', 'fat_index', 'odc_index', 'pole_count', 'headpole_count', 'hook_count'],
                'باکس دانلود': ['documents_box', 'download_form', 'documents','marketer_name'],
                'اطلاعات شخصی': ['first_name', 'last_name', 'father_name', 'national_code', 'originated_from', 'bc_number', 'birthday', 'landline_number', 'mobile_number', 'address', 'location', 'post_code', 'house_is_owner'],
                'اطلاعات سرویس': ['sip_phone', 'modem', 'plan'],
                'سایر اطلاعات': ['finished_request', 'contact_user', 'tracking_code', 'log_msg_status', 'jalali_request_time'],
            },
            'role_dropagent': {
                'وضعیت نصب': ['drop_status'],
                'اطلاعات دراپ': ['outdoor_area', 'internal_area', 'fat_index', 'odc_index', 'pole_count', 'headpole_count', 'hook_count'],
                'اطلاعات شخصی': ['first_name', 'last_name', 'landline_number', 'mobile_number', 'address', 'location', 'post_code'],
                'سایر اطلاعات': ['contact_user', 'tracking_code', 'jalali_request_time'],
            },
            'role_supervisor': {
                'وضعیت نصب': ['supervisor_status'],
                'اطلاعات دراپ': ['outdoor_area', 'internal_area', 'fat_index', 'odc_index', 'pole_count', 'headpole_count', 'hook_count'],
                'اطلاعات شخصی': ['first_name', 'last_name', 'landline_number', 'mobile_number', 'address', 'location', 'post_code'],
                'سایر اطلاعات': ['contact_user', 'tracking_code', 'jalali_request_time'],
            },
            'role_fusionagent': {
                'وضعیت نصب': ['fusion_status'],
                'وضعیت پرداخت':['pay_status', 'account_number', 'tracking_payment','payment_date','payment_time'],
                'اطلاعات دراپ': ['outdoor_area', 'internal_area', 'fat_index', 'odc_index', 'pole_count', 'headpole_count', 'hook_count'],
                'اطلاعات شخصی': ['first_name', 'last_name', 'landline_number', 'mobile_number', 'address', 'location', 'post_code'],
                'سایر اطلاعات': ['contact_user', 'tracking_code', 'jalali_request_time'],
            },
            'role_operator': {
                'وضعیت نصب': ['submission_status', 'virtual_number','port_number'],
                'اطلاعات دراپ': ['outdoor_area', 'internal_area', 'fat_index', 'odc_index', 'pole_count', 'headpole_count', 'hook_count'],
                'باکس دانلود': ['documents_box', 'download_form', 'documents','marketer_name'],
                'اطلاعات شخصی': ['first_name', 'last_name', 'father_name', 'national_code', 'originated_from', 'bc_number', 'birthday', 'landline_number', 'mobile_number', 'address', 'location', 'post_code', 'house_is_owner'],
                'سایر اطلاعات': ['contact_user', 'tracking_code', 'jalali_request_time'],
            },
        }
        
        # --- ۳. ساخت مجموعه (Set) کل فیلدهای مجاز برای کاربر ---
        # اینجا از set استفاده می‌کنیم چون ترتیب مهم نیست و فقط می‌خواهیم دسترسی را چک کنیم
        all_allowed_fields = set()

        if getattr(request.user, "role_marketer", False):
            for section, fields in ROLE_FIELDS['role_marketer'].items():
                all_allowed_fields.update(fields)

        if getattr(request.user, "role_dropagent", False):
            for section, fields in ROLE_FIELDS['role_dropagent'].items():
                all_allowed_fields.update(fields)

        if getattr(request.user, "role_supervisor", False):
            for section, fields in ROLE_FIELDS['role_supervisor'].items():
                all_allowed_fields.update(fields)

        if getattr(request.user, "role_fusionagent", False):
            for section, fields in ROLE_FIELDS['role_fusionagent'].items():
                all_allowed_fields.update(fields)

        if getattr(request.user, "role_operator", False):
            for section, fields in ROLE_FIELDS['role_operator'].items():
                all_allowed_fields.update(fields)

        # --- ۴. ساخت fieldsets نهایی با حفظ ترتیب ---
        final_fieldsets = []
        # روی ساختار مرجع (Master) که ترتیب درست دارد، حلقه می‌زنیم
        for section_title, options in MASTER_FIELDSETS_STRUCTURE:
            master_fields_in_section = options.get('fields', ())
            
            # فیلدها را بر اساس لیست مرجع فیلتر می‌کنیم
            visible_fields_for_this_section = []
            for field_name in master_fields_in_section:
                # اگر فیلد در لیست مرجع، در مجموعه فیلدهای مجاز کاربر بود...
                if field_name in all_allowed_fields:
                    visible_fields_for_this_section.append(field_name) # ...به لیست نهایی اضافه کن
            
            # فقط در صورتی سکشن را اضافه کن که فیلدی برای نمایش داشته باشد
            if visible_fields_for_this_section:
                final_fieldsets.append((
                    section_title,
                    {
                        # اینجا یک تاپل مرتب شده بر اساس لیست مرجع داریم
                        'fields': tuple(visible_fields_for_this_section), 
                        'classes': options.get('classes', ('collapse',))
                    }
                ))
        
        return tuple(final_fieldsets)

    def get_readonly_fields(self, request, obj=None):
        if obj:
            request.session['secure_form_download'] = obj.pk

        method_fields = {
            "contact_user",
            "download_form",
            "documents_box",
            "documents_upload",
            "view_contract_button", 
            "log_msg_status",
            "ip_address",
            "jalali_request_time",
        }

        if request.user.is_superuser:
            return list(method_fields)
        
        WRITABLE_FIELDS_BY_ROLE = {
            'role_marketer': {
                'marketer_status',
                'pay_status',
                'account_number',
                'tracking_payment',
                'payment_date',
                'payment_time',
                'documents',
                'marketer_name',
                'first_name',
                'last_name', 
                'father_name',
                'national_code',
                'originated_from', 
                'bc_number',
                'birthday', 
                'landline_number', 
                'mobile_number',
                'address',
                'location', 
                'post_code',
                'house_is_owner',
                'sip_phone',
                'modem',
                'plan',
                'finished_request',
            },
            'role_dropagent': {
                'drop_status',
                'outdoor_area',
                'internal_area',
                'fat_index',
                'odc_index',
                'pole_count',
                'headpole_count',
                'hook_count',
            },
            'role_supervisor': {
                'supervisor_status',
            },
            'role_fusionagent': {
                'fusion_status',
                'pay_status',
                'account_number',
                'tracking_payment',
                'payment_date',
                'payment_time',
            },
            'role_operator': {
                'submission_status',
                'virtual_number',
                'port_number',
            }
        }
        

        all_visible_fields = set()
        fieldsets = self.get_fieldsets(request, obj)
        for section_name, options in fieldsets:
            if 'fields' in options:
                all_visible_fields.update(options['fields'])


        user_writable_fields = set()
        
        if getattr(request.user, "role_marketer", False):
            user_writable_fields.update(WRITABLE_FIELDS_BY_ROLE.get('role_marketer', set()))

        if getattr(request.user, "role_dropagent", False):
            user_writable_fields.update(WRITABLE_FIELDS_BY_ROLE.get('role_dropagent', set()))

        if getattr(request.user, "role_supervisor", False):
            user_writable_fields.update(WRITABLE_FIELDS_BY_ROLE.get('role_supervisor', set()))

        if getattr(request.user, "role_fusionagent", False):
            user_writable_fields.update(WRITABLE_FIELDS_BY_ROLE.get('role_fusionagent', set()))

        if getattr(request.user, "role_operator", False):
            user_writable_fields.update(WRITABLE_FIELDS_BY_ROLE.get('role_operator', set()))

        # ۳. محاسبه فیلدهای Readonly
        # (تمام فیلدهای قابل مشاهده) - (فیلدهای قابل ویرایش)
        role_based_readonly_fields = all_visible_fields - user_writable_fields
        
        # ۴. ادغام لیست متدها با لیست readonly مبتنی بر نقش
        # این تضمین می‌کند که متدها همیشه readonly باقی بمانند
        final_readonly_fields = role_based_readonly_fields.union(method_fields)
        
        return list(final_readonly_fields)

    def get_form(self, request, obj=None, **kwargs): 

        form = super().get_form(request, obj, **kwargs)
        if obj:
            obj.refresh_from_db()

            try:
                if not request.user.is_superuser:
                    if obj.marketer_name:
                        form.base_fields['marketer_name'].disabled = True
                        form.base_fields['marketer_name'].widget = DisplayOnlyWidget(obj.get_full_name(),color="#B6771D")
                
            #رد انلی شدن تایید مدارک
                if getattr(request.user, "role_marketer", False) or request.user.is_superuser:
                    if obj.drop_status == "accepted":

                        form.base_fields['marketer_status'].disabled = True
                        form.base_fields['marketer_status'].widget = DisplayOnlyWidget(
                        "امکان تغییر این بخش وجود ندارد ، مدارک تایید و دراپ کشی انجام شده است", color="#B6771D")
                
                    if obj.drop_status == "queued" :
            
                        form.base_fields['marketer_status'].disabled = True
                        form.base_fields['marketer_status'].widget = DisplayOnlyWidget(
                        "امکان تغییر این بخش وجود ندارد ، سرویس در صف دراپ کشی قرار گرفته است", color="#B6771D")
        
                #رد انلی شدن دراپ 
                if getattr(request.user, "role_dropagent", False) or request.user.is_superuser:
                    if obj.marketer_status != "accepted":
                        form.base_fields['drop_status'].disabled = True
                        form.base_fields['drop_status'].widget = DisplayOnlyWidget(
                        "در انتظار تایید مدارک و اطلاعات توسط بازاریاب", color="crimson"
                            )
                    if obj.fusion_status != "pending" :
                        form.base_fields['drop_status'].disabled = True
                        form.base_fields['drop_status'].widget = DisplayOnlyWidget(
                        "امکان تغییر وضعیت دراپ وجود ندارد ، سرویس در حال فیوژن زنی است", color="#B6771D"
                            )
                #رد انلی شدن فیوژن
                if getattr(request.user, "role_fusionagent", False) or request.user.is_superuser:
                    if obj.drop_status != "accepted":
                        form.base_fields['fusion_status'].disabled = True
                        form.base_fields['fusion_status'].widget = DisplayOnlyWidget(
                        "امکان فیوژن زنی وجود ندارد دراپ کشی هنوز انجام نشده است", color="crimson")
                    elif obj.supervisor_status == "accepted" :
                        form.base_fields['fusion_status'].disabled = True
                        form.base_fields['fusion_status'].widget = DisplayOnlyWidget(
                        "امکان تغییر وجود ندارد ، سرویس توسط ناظر تایید شده است", color="#B6771D")

                #رد انلی شدن ناظر
                if getattr(request.user, "role_supervisor", False) or request.user.is_superuser:
                    if obj.fusion_status != "accepted" :
                        form.base_fields['supervisor_status'].disabled = True
                        form.base_fields['supervisor_status'].widget = DisplayOnlyWidget(
                        "امکان بازبینی وجود ندارد ، دراپ و فیوژن هنوز اجرانشده است", color="crimson")                     
                                        
                    elif obj.submission_status == "registered":
                        form.base_fields['supervisor_status'].disabled = True
                        form.base_fields['supervisor_status'].widget = DisplayOnlyWidget(
                        "امکان تغییر وجود ندارد ، فرم ثبت شده است", color="#B6771D")      
                        
                #رد انلی شدن ثبت فرم
                if getattr(request.user, "role_operator", False) or request.user.is_superuser :
                    if obj.supervisor_status != "accepted":
                        form.base_fields['submission_status'].disabled = True
                        form.base_fields['submission_status'].widget = DisplayOnlyWidget(
                        "امکان ثبت فرم وجود ندارد ، ناظر هنوز نصب را تایید نکرده است", color="crimson")
                    elif obj.finalization_status == "ended":

                        form.base_fields['virtual_number'].disabled = True
                        form.base_fields['virtual_number'].widget = DisplayOnlyWidget(
                            obj.virtual_number,color="#B6771D"
                        )
                        form.base_fields['port_number'].disabled = True
                        form.base_fields['port_number'].widget = DisplayOnlyWidget(
                            obj.port_number,color="#B6771D"
                        )

                        form.base_fields['submission_status'].disabled = True
                        form.base_fields['submission_status'].widget = DisplayOnlyWidget(
                        "امکان تغییر وضعیت وجود ندارد ، سرویس ثبت نهایی شده است", color="#B6771D")

                # رد انلی شدن فانالیزیشن
                if getattr(request.user, "role_fusionagent", False) or request.user.is_superuser:
                    if obj.submission_status != "registered":
                        form.base_fields['finalization_status'].disabled = True
                        form.base_fields['finalization_status'].widget = DisplayOnlyWidget(
                        "امکان ثبت نهایی وجود ندارد فرم هنوز ثبت نشده است", color="crimson")                     
 

            except Exception as e:
                print(e)
     
        return form
    
    actions = ['export_excel_information','generate_snapshots']

    @admin.action(description='خروجی اکسل برای کاربران انتخاب شده')
    def export_excel_information(self, request, queryset):
        if (request.user.is_superuser != True and request.user.role_marketer != True) :
            self.message_user(request, "شما دسترسی لازم برای انجام این کار را ندارید.", level='error')
            return
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.sheet_view.rightToLeft = True

        # --- مرحله ۲: نوشتن تایتل اصلی (ردیف ۱) ---
        title = "لیست بازاریابی گروه توسعه زیر ساخت نوراجم"
        # تعریف هدرها برای پیدا کردن تعداد ستون‌ها
        headers = [
            "ردیف", "نام و نام خانوادگی", "کد پستی", "کد ملی", "شماره تماس",
            "آدرس پستی", "مشخصه FAT", "تعداد تیر","تعداد سرتیر", "وضعیت پرداخت",
            "وضعیت دراپ", "وضعیت فیوژن", "وضعیت ثبت نام", "متراژ خارجی",
            "متراژ داخلی", "نام بازاریاب", "تاریخ درخواست"
        ]
        num_columns = len(headers)
        
        # ادغام سلول‌های ردیف اول برای تایتل
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # --- مرحله ۳: نوشتن هدر ستون‌ها (ردیف ۲) ---
        ws.append(headers)
        # استایل‌دهی به ردیف هدر
        for col_num in range(1, num_columns + 1):
            cell = ws.cell(row=2, column=col_num)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='right', vertical='center')

        # --- مرحله ۴: نوشتن داده‌ها (از ردیف ۳ به بعد) ---
        row_counter = 1
        for obj in queryset:
            # مدیریت زمان (برای نمایش خوانا و محلی)
            request_time_str = ""
            if obj.request_time:
                request_time_str = to_jalali(obj.request_time)

            # ساخت ردیف داده‌ها با مدیریت مقادیر Null (تبدیل به رشته خالی)
            data_row = [
                f"{row_counter:02d}",  # ردیف با فرمت 01, 02
                obj.get_full_name(),
                obj.post_code or "",
                obj.national_code or "",
                obj.mobile_number or "",
                obj.address or "",
                obj.fat_index or "",
                obj.pole_count or "",
                obj.headpole_count or "",
                obj.get_pay_status_display(),
                obj.get_drop_status_display(),
                obj.get_fusion_status_display(),
                obj.get_marketer_status_display(),
                obj.outdoor_area or "",
                obj.internal_area or "",
                "" if not obj.marketer_name else obj.marketer_name.get_full_name(),
                request_time_str,
            ]
            
            ws.append(data_row)
            
            # استایل‌دهی به سلول‌های داده (اختیاری، برای خوانایی)
            for col_num in range(1, num_columns + 1):
                ws.cell(row=row_counter + 2, column=col_num).alignment = Alignment(horizontal='right', vertical='center')
                
            row_counter += 1

        # --- مرحله ۵: تنظیم عرض ستون‌ها (اختیاری اما بسیار مفید) ---
        dim_holder = DimensionHolder(worksheet=ws)
        for col in range(1, num_columns + 1):
            # یک عرض پیش‌فرض تعیین می‌کنیم
            width = 20
            if col == 1: # ردیف
                width = 5
            elif col == 6: # آدرس پستی
                width = 40
            elif col == 2: # نام
                width = 25
                
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=width)
        
        ws.column_dimensions = dim_holder

        # --- مرحله ۶: ایجاد و بازگرداندن پاسخ HTTP ---
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="marketing_export.xlsx"'
        
        # ذخیره ورک‌بوک در پاسخ
        wb.save(response)
        
        return response


    @admin.action(description="ساخت اسنپ‌شات قرارداد برای موارد انتخاب شده")
    def generate_snapshots(self, request, queryset):
        if (request.user.is_superuser != True) :
            self.message_user(request, "شما دسترسی لازم برای انجام این کار را ندارید.", level='error')
            return
        updated_count = 0
        skipped_count = 0

        target_queryset = queryset.filter(
            plan__isnull=False, 
            modem__isnull=False,
            sip_phone__isnull=False,
        )

        for obj in target_queryset:
            try:
                obj.contract_snapshot = obj.generate_contract_snapshot()
                obj.save(update_fields=['contract_snapshot'])
                updated_count += 1
            except Exception as e:
                print(e)
        
        skipped_count = queryset.count() - updated_count

        if updated_count > 0:
            self.message_user(request, f"{updated_count} اسنپ‌شات قرارداد با موفقیت ساخته شد.", messages.SUCCESS)
        if skipped_count > 0:
            self.message_user(request, f"{skipped_count} مورد نادیده گرفته شد (احتمالاً از قبل اسنپ‌شات داشتند یا اطلاعات ناقص بود).", messages.WARNING)

@admin.register(ActiveLocations,site=super_admin_site)
class ActiveLocationsAdmin(admin.ModelAdmin):
    list_display = ("name","area_limit","is_active")

    def has_view_permission(self, request, obj=None):
        return (True if request.user.is_superuser or request.user.role_marketer else False)

    def has_add_permission(self, request, obj=None):
        return (True if request.user.is_superuser or request.user.role_marketer else False)

    def has_change_permission(self, request, obj=None):
        return (True if request.user.is_superuser or request.user.role_marketer else False)
    
    def has_delete_permission(self, request, obj=None):
        return (True if request.user.is_superuser or request.user.role_marketer else False)

@admin.register(ActiveModems,site=super_admin_site)
class ActiveModemsAdmin(admin.ModelAdmin):
    list_display = ("name","price","payment_method","is_active")

    def has_view_permission(self, request, obj=None):
        return (True if request.user.is_superuser or request.user.role_marketer else False)

    def has_add_permission(self, request, obj=None):
        return (True if request.user.is_superuser or request.user.role_marketer else False)

    def has_change_permission(self, request, obj=None):
        return (True if request.user.is_superuser or request.user.role_marketer else False)
    
    def has_delete_permission(self, request, obj=None):
        return (True if request.user.is_superuser or request.user.role_marketer else False)

@admin.register(ActivePlans,site=super_admin_site)
class ActivePlansAdmin(admin.ModelAdmin):
    list_display = ("data","price","is_active")
    ordering = ("data","is_active","price")

    def has_view_permission(self, request, obj=None):
        return (True if request.user.is_superuser or request.user.role_marketer else False)

    def has_add_permission(self, request, obj=None):
        return (True if request.user.is_superuser or request.user.role_marketer else False)

    def has_change_permission(self, request, obj=None):
        return (True if request.user.is_superuser or request.user.role_marketer else False)
    
    def has_delete_permission(self, request, obj=None):
        return (True if request.user.is_superuser or request.user.role_marketer else False)

class AccountNumberInline(admin.TabularInline):
    model = AccountNumber
    extra = 1
    verbose_name_plural = "شماره های حساب"

@admin.register(OtherInfo,site=super_admin_site)
class OtherInfoAdmin(admin.ModelAdmin):
    list_display = ("sip_phone_cost","drop_cost","center_name","center_address")

    inlines = [AccountNumberInline]

    def has_view_permission(self, request, obj=None):
        return (True if request.user.is_superuser or request.user.role_marketer else False)

    def has_add_permission(self, request, obj=None):
        return (True if request.user.is_superuser or request.user.role_marketer else False)

    def has_change_permission(self, request, obj=None):
        return (True if request.user.is_superuser or request.user.role_marketer else False)
    
    def has_delete_permission(self, request, obj=None):
        return (True if request.user.is_superuser or request.user.role_marketer else False)

User = get_user_model()

class UserAdminForm(UserChangeForm):

    class Meta(UserChangeForm.Meta):
        model = User

@admin.register(User,site=super_admin_site)
class UserAdmin(BaseUserAdmin):
    form = UserAdminForm

    list_display_links = ("full_name",)

    list_display = ("full_name", "username","role")

    ordering = ("-date_joined",)

    readonly_fields = ("jalali_last_login", "jalali_date_joined")

    list_filter = ()

    fieldsets = (
        (None, {'fields': ('username', 'password')}),
        ('اطلاعات شخصی', {'fields': ('first_name', 'last_name', 'email')}),
        ('نقش ها',{'fields':('role_supervisor','role_marketer','role_dropagent','role_fusionagent','role_operator')}),
        # ('دسترسی ها', {'fields': ('is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions')}),
        ('دسترسی ها', {'fields': ('is_active', 'is_staff', 'is_superuser', 'user_permissions','register_messages')}), # گروه ها حذف شد 
        ('تاریخ های مهم', {'fields': ('jalali_last_login', 'jalali_date_joined')}),
    )

    def has_add_permission(self, request): return request.user.is_superuser

    def has_delete_permission(self, request,obj=None): return request.user.is_superuser

    def get_fieldsets(self, request, obj=None): # type: ignore
        if not request.user.is_superuser:
            return (
        ('اطلاعات پایه', {'fields': ('username', 'password')}),
        ('اطلاعات شخصی', {'fields': ('first_name', 'last_name', 'email')}),

    )
        return super().get_fieldsets(request, obj)

    def get_queryset(self, request): # فقط خودش رو مشاهده کنه 
        self._current_user = request.user  
        self._full_name_user = (request.user.first_name)+" "+(request.user.last_name)
        queryset = super().get_queryset(request)
        if not request.user.is_superuser:
            return queryset.filter(pk=request.user.pk)
        return queryset

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        if not change and not obj.is_superuser:
            # همه دسترسی‌ها رو بده به کاربر
            obj.user_permissions.set(Permission.objects.all())
            obj.is_staff = True
            obj.save()

    def jalali_last_login(self, obj):
        return to_jalali_persian(obj.last_login)
    jalali_last_login.short_description = "آخرین ورود"
    jalali_last_login.admin_order_field = 'last_login'

    def jalali_date_joined(self, obj):
        return to_jalali_persian(obj.date_joined)
    jalali_date_joined.short_description = "تاریخ عضویت"
    jalali_date_joined.admin_order_field = 'date_joined'

    def full_name(self,obj):
        return obj.first_name+" "+obj.last_name
    full_name.short_description = "نام"

    def role(self,obj):
        return format_html(f'<span style="color:var(--link-fg);">{obj.get_role()}</span>')
    role.short_description = "نقش"
        
    actions = ['send_custom_sms']

    @admin.action(description='ارسال پیامک به کاربران انتخاب شده')
    def send_custom_sms(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request, "شما دسترسی لازم برای انجام این کار را ندارید.", level='error')
            return
        
        selected_ids = list(queryset.values_list('id', flat=True))
        request.session['selected_user_ids_for_sms'] = selected_ids

        return HttpResponseRedirect(reverse('send_sms_page'))
    
